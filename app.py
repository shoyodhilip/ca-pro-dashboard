import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, date
import matplotlib.pyplot as plt
import os

# ---------------- SETTINGS ---------------- #
FILES = {
    "FR": "FR.xlsx",
    "AFM": "AFM.xlsx",
    "Audit": "Audit.xlsx",
    "DT": "DT.xlsx",
    "IDT": "IDT.xlsx"
}

EXAM_DATE = datetime(2026, 8, 30)

# ---------------- PAGE CONFIG ---------------- #
st.set_page_config(
    page_title="CA Final Pro Tracker",
    layout="wide"
)

# ---------------- MODERN UI ---------------- #
st.markdown("""
<style>

.stApp {
    background-color: #0E1117;
    color: white;
}

h1, h2, h3 {
    color: white;
}

div[data-testid="metric-container"] {
    background: #1E1E2F;
    padding: 20px;
    border-radius: 15px;
    border: 1px solid #333;
    text-align: center;
}

div.stButton > button {
    width: 100%;
    border-radius: 10px;
    height: 3em;
    background-color: #4CAF50;
    color: white;
    border: none;
    font-size: 16px;
}

.stCheckbox {
    background-color: #1E1E2F;
    padding: 10px;
    border-radius: 10px;
    margin-bottom: 5px;
}

</style>
""", unsafe_allow_html=True)

# ---------------- TIME FIX ---------------- #
def convert_to_hours(duration):

    if duration is None:
        return 0

    try:

        if isinstance(duration, (int, float)):

            if pd.isna(duration):
                return 0

            return float(duration)

        if hasattr(duration, "hour"):
            return duration.hour + duration.minute / 60

        if hasattr(duration, "total_seconds"):
            return duration.total_seconds() / 3600

    except:
        return 0

    return 0

# ---------------- HEADER ---------------- #
st.title("📚 CA Final Pro Dashboard")

# ---------------- DASHBOARD ---------------- #
total_hours = 0
completed_hours = 0

subject_data = {}

for subject, file in FILES.items():

    if not os.path.exists(file):
        continue

    wb = load_workbook(file)
    ws = wb.active

    total = 0
    done = 0

    for row in ws.iter_rows(min_row=2, values_only=True):

        duration = row[3]
        completed = row[4]

        hrs = convert_to_hours(duration)

        total += hrs

        if completed:
            done += hrs

    subject_data[subject] = (done, total)

    total_hours += total
    completed_hours += done

# ---------------- COUNTDOWN ---------------- #
days_left = (EXAM_DATE.date() - date.today()).days

remaining = total_hours - completed_hours

required_per_day = remaining / days_left if days_left else 0

col1, col2, col3 = st.columns(3)

col1.metric("⏳ Days Left", days_left)
col2.metric("📖 Completed Hours", round(completed_hours, 1))
col3.metric("🔥 Required / Day", round(required_per_day, 1))

# ---------------- OVERALL PROGRESS ---------------- #
overall_percent = (completed_hours / total_hours * 100) if total_hours else 0

st.subheader("📊 Overall Progress")

st.progress(overall_percent / 100)

st.write(f"### {overall_percent:.1f}% Completed")

# ---------------- SUBJECT SELECT ---------------- #
st.subheader("📘 Subject Tracker")

selected_subject = st.selectbox(
    "Choose Subject",
    ["Select"] + list(FILES.keys()),
    index=0
)

# ---------------- SUBJECT TOPICS ---------------- #
if selected_subject != "Select":

    file = FILES[selected_subject]

    if os.path.exists(file):

        wb = load_workbook(file)
        ws = wb.active

        st.write(f"## {selected_subject}")

        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):

            day = row[0].value
            part = row[1].value
            topic = row[2].value
            duration = row[3].value
            completed = row[4].value

            hrs = convert_to_hours(duration)

            checkbox = st.checkbox(
                f"Day {day} | Part {part} | {topic} ({hrs:.1f} hrs)",
                value=bool(completed),
                key=f"{selected_subject}_{i}"
            )

            # SAVE TO EXCEL
            if checkbox != bool(completed):

                ws.cell(row=i, column=5).value = checkbox

                if checkbox:
                    ws.cell(row=i, column=6).value = str(date.today())
                else:
                    ws.cell(row=i, column=6).value = ""

                wb.save(file)

                st.success("Progress Updated ✅")

                st.rerun()

# ---------------- SUBJECT PROGRESS ---------------- #
st.subheader("📈 Subject Progress")

for subject, values in subject_data.items():

    done, total = values

    percent = (done / total * 100) if total else 0

    st.write(f"### {subject}")

    st.progress(percent / 100)

    st.write(f"{done:.1f}/{total:.1f} hrs")

# ---------------- ADVANCED CHARTS ---------------- #
st.subheader("📊 Study Analytics")

remaining = max(total_hours - completed_hours, 0)

if total_hours > 0:

    col1, col2 = st.columns(2)

    # PIE CHART
    with col1:

        labels = ["Completed", "Remaining"]
        sizes = [completed_hours, remaining]

        fig1, ax1 = plt.subplots()

        ax1.pie(
            sizes,
            labels=labels,
            autopct='%1.1f%%'
        )

        ax1.set_title("Overall Completion")

        st.pyplot(fig1)

    # BAR CHART
    with col2:

        subjects = list(subject_data.keys())

        percentages = []

        for s in subjects:

            done, total = subject_data[s]

            percent = (done / total * 100) if total else 0

            percentages.append(percent)

        fig2, ax2 = plt.subplots()

        ax2.bar(subjects, percentages)

        ax2.set_ylabel("Completion %")

        ax2.set_title("Subject Progress")

        st.pyplot(fig2)

else:
    st.warning("No study data available yet.")

# ---------------- FOOTER ---------------- #
st.markdown("---")
st.write("🔥 CA Final Success Dashboard")
